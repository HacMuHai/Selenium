"""
ImportService - đọc Excel đã export và dựng lại document product để nạp vào Mongo.

Đường ngược của `ExportService`: `data/<sàn>/comments_export_*.xlsx` -> collection Mongo.
Dùng khi crawl ở chế độ `--no-db` (Excel là nguồn dữ liệu duy nhất) và sau này mới có DB.

Ba điểm cần biết về định dạng Excel:
- Cột đọc theo TÊN trong dòng header, không theo vị trí, nên file cũ 4 cột vẫn nạp được.
- `link` và `name_item` chỉ điền ở dòng đầu mỗi product, các dòng sau để trống -> phải
  forward-fill theo thứ tự dòng. File cũ 4 cột không có cột `site` thì suy từ hostname.
- Một product có thể trải trên nhiều file (bị cắt theo `max_rows_per_file`), nên gom
  theo `link` trên TOÀN BỘ thư mục rồi mới ghi, không ghi theo từng file.

Trùng lặp: comment gom theo `comments_id`; product đã có trong DB thì tuỳ `mode`
(`skip` bỏ qua, `replace` ghi đè cả document).
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Iterator, Optional

from openpyxl import load_workbook

from src.config import version as version_module
from src.services.export_service import UNKNOWN_SITE, site_of

logger = logging.getLogger(__name__)

COL_LINK = "link"
COL_NAME = "name_item"
COL_ID = "comments_id"
COL_CONTENT = "comments_content"
COL_SITE = "site"
COL_USER = "user_name"
COL_RATING = "rating"


def iter_excel_files(input_dir: str) -> list[Path]:
    """Mọi .xlsx trong thư mục và thư mục con (bỏ file tạm ~$ của Excel)."""
    directory = Path(input_dir)
    if not directory.is_dir():
        raise FileNotFoundError(f"Không thấy thư mục đầu vào: {directory}")
    return sorted(p for p in directory.rglob("*.xlsx") if not p.name.startswith("~$"))


def _cell(row: tuple, idx: Optional[int]) -> Optional[str]:
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    text = str(row[idx]).strip()
    return text or None


def read_products(paths: list[Path]) -> dict[str, dict]:
    """Đọc nhiều file Excel, trả `{link: product}` đã gom comment và khử trùng lặp."""
    products: dict[str, dict] = {}

    for path in paths:
        try:
            sheet = load_workbook(path, read_only=True).active
            rows = sheet.iter_rows(values_only=True)
            header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
        except (StopIteration, OSError, KeyError):
            logger.warning("Bỏ qua file không đọc được: %s", path, exc_info=True)
            continue

        if COL_CONTENT not in header:
            logger.warning("Bỏ qua %s: thiếu cột %s", path, COL_CONTENT)
            continue

        idx = {name: (header.index(name) if name in header else None) for name in (
            COL_LINK, COL_NAME, COL_ID, COL_CONTENT, COL_SITE, COL_USER, COL_RATING
        )}

        current_link: Optional[str] = None
        current_name = ""
        rows_read = 0

        for row in rows:
            # Forward-fill: dòng trống 2 cột đầu thuộc về product của dòng trước.
            link = _cell(row, idx[COL_LINK])
            if link:
                current_link = link
                current_name = _cell(row, idx[COL_NAME]) or ""
            if current_link is None:
                continue  # dòng mồ côi trước product đầu tiên

            content = _cell(row, idx[COL_CONTENT])
            if not content:
                continue

            product = products.get(current_link)
            if product is None:
                product = {
                    "name": current_name,
                    "link": current_link,
                    "site": _cell(row, idx[COL_SITE])
                    or site_of({"link": current_link})
                    or UNKNOWN_SITE,
                    "comments": [],
                    "_seen_ids": set(),
                }
                products[current_link] = product
            elif current_name and not product["name"]:
                product["name"] = current_name

            comment_id = _cell(row, idx[COL_ID]) or f"{current_link}#{len(product['comments'])}"
            if comment_id in product["_seen_ids"]:
                continue
            product["_seen_ids"].add(comment_id)
            product["comments"].append(
                {
                    "id": comment_id,
                    "name": _cell(row, idx[COL_USER]) or "",
                    "content": content,
                    "rating": _to_rating(_cell(row, idx[COL_RATING])),
                }
            )
            rows_read += 1

        logger.info("%s: %d dòng comment", path, rows_read)

    return products


def _to_rating(value: Optional[str]) -> int:
    """Cột rating có thể trống (file 4 cột cũ) hoặc là số thực -> ép về int, lỗi thì 0."""
    if value is None:
        return 0
    try:
        return int(float(value))
    except ValueError:
        return 0


class ImportService:
    """Nạp Excel vào repository. `mode`: `skip` giữ product đã có, `replace` ghi đè."""

    def __init__(self, repository, mode: str = "skip") -> None:
        if mode not in ("skip", "replace"):
            raise ValueError("mode phải là 'skip' hoặc 'replace'")
        self.repository = repository
        self.mode = mode

    def import_dir(self, input_dir: str, dry_run: bool = False) -> dict:
        paths = iter_excel_files(input_dir)
        logger.info("Tìm thấy %d file Excel trong %s", len(paths), input_dir)
        products = read_products(paths)

        stats = {"files": len(paths), "products": len(products), "inserted": 0,
                 "replaced": 0, "skipped": 0, "comments": 0}
        now = datetime.now()

        for link, product in products.items():
            product.pop("_seen_ids", None)
            comments = product["comments"]
            stats["comments"] += len(comments)
            document = {
                **product,
                "total_comments": len(comments),
                "crawled_at": now,
                "version": version_module.version,
            }

            exists = self.repository.exists_by_link(link)
            if exists and self.mode == "skip":
                stats["skipped"] += 1
                continue
            if dry_run:
                stats["replaced" if exists else "inserted"] += 1
                continue
            if exists:
                self.repository.replace_by_link(link, document)
                stats["replaced"] += 1
            else:
                self.repository.insert_product(document)
                stats["inserted"] += 1

        logger.info(
            "Import%s: %d file, %d product (%d thêm, %d ghi đè, %d bỏ qua), %d comment",
            " [dry-run]" if dry_run else "",
            stats["files"], stats["products"], stats["inserted"],
            stats["replaced"], stats["skipped"], stats["comments"],
        )
        return stats
