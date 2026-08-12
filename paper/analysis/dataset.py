"""
Đọc dữ liệu đã gán nhãn từ `data_tagged/`, làm sạch, chia train/test.

RỦI RO SỐ 1 CỦA MODULE NÀY LÀ RÒ RỈ DỮ LIỆU: cùng một câu comment xuất hiện nhiều lần
trong dữ liệu. Nếu split trước khi khử trùng, câu đó nằm ở CẢ train lẫn test → điểm số
ảo cao → toàn bộ so sánh 3 model vô nghĩa. **Khử trùng TRƯỚC, split SAU.**
"""
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator, Optional

import pandas as pd
from openpyxl import load_workbook

from paper.analysis.preprocessing import normalize_text

logger = logging.getLogger(__name__)

LABELS = ("negative", "neutral", "positive")

COL_ID = "comments_id"
COL_TEXT = "comments_content"
COL_TAG = "tag"
COL_LINK = "link"
COL_NAME = "name_item"


@dataclass
class CleaningStats:
    total_rows: int = 0
    dropped_empty: int = 0
    dropped_bad_label: int = 0
    dropped_conflict: int = 0
    dropped_duplicate: int = 0
    skipped_files: list[str] = field(default_factory=list)
    final_rows: int = 0
    label_counts: dict[str, int] = field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            "total_rows": self.total_rows,
            "dropped_empty": self.dropped_empty,
            "dropped_bad_label": self.dropped_bad_label,
            "dropped_conflict": self.dropped_conflict,
            "dropped_duplicate": self.dropped_duplicate,
            "skipped_files": self.skipped_files,
            "final_rows": self.final_rows,
            "label_counts": self.label_counts,
        }


def _column_index(header: list[str], name: str) -> int:
    """Vị trí cột theo TÊN, không hardcode chỉ số.

    `.index()` lấy cột đầu tiên trùng tên → file có `tag` lặp nhiều lần vẫn đọc đúng,
    và không nhầm sang `tag_cu` / `nguon_tag`.
    """
    return header.index(name)


def _read_file(path: Path, stats: CleaningStats) -> list[dict]:
    """Đọc 1 file Excel đã gán nhãn. File hỏng/thiếu cột → bỏ qua cả file, không giết pipeline."""
    try:
        sheet = load_workbook(path, read_only=True).active
        rows = sheet.iter_rows(values_only=True)
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
        idx_id = _column_index(header, COL_ID)
        idx_text = _column_index(header, COL_TEXT)
        idx_tag = _column_index(header, COL_TAG)
    except (StopIteration, ValueError, OSError, KeyError):
        logger.warning("Bỏ qua file không đọc được / thiếu cột: %s", path.name, exc_info=True)
        stats.skipped_files.append(path.name)
        return []

    out: list[dict] = []
    for row in rows:
        stats.total_rows += 1
        raw_text = row[idx_text] if idx_text < len(row) else None
        norm = normalize_text(raw_text)
        if not norm:
            stats.dropped_empty += 1
            continue

        raw_tag = row[idx_tag] if idx_tag < len(row) else None
        label = str(raw_tag).strip().lower() if raw_tag is not None else ""
        if label not in LABELS:
            stats.dropped_bad_label += 1
            continue

        out.append(
            {
                "comment_id": str(row[idx_id]) if idx_id < len(row) and row[idx_id] else "",
                "text": str(raw_text),
                "norm_text": norm,
                "sentiment": label,
            }
        )
    return out


def load_tagged_dataset(tag_dir: str) -> tuple[pd.DataFrame, CleaningStats]:
    """Gom mọi file `.xlsx` trong `tag_dir` → DataFrame đã khử trùng, không mâu thuẫn nhãn."""
    directory = Path(tag_dir).resolve()
    if not directory.is_dir():
        raise FileNotFoundError(f"Không thấy thư mục dữ liệu đã gán nhãn: {directory}")

    stats = CleaningStats()
    records: list[dict] = []
    for path in sorted(directory.glob("*.xlsx")):
        records.extend(_read_file(path, stats))

    if not records:
        raise ValueError(f"Không đọc được dòng hợp lệ nào từ {directory}")

    frame = pd.DataFrame.from_records(records)

    # Cùng nội dung nhưng >1 nhãn -> bỏ CẢ NHÓM, không đoán bừa nhãn nào đúng.
    label_per_text = frame.groupby("norm_text")["sentiment"].nunique()
    conflicting = set(label_per_text[label_per_text > 1].index)
    if conflicting:
        stats.dropped_conflict = int(frame["norm_text"].isin(conflicting).sum())
        frame = frame[~frame["norm_text"].isin(conflicting)]

    before = len(frame)
    frame = frame.drop_duplicates(subset="norm_text", keep="first").reset_index(drop=True)
    stats.dropped_duplicate = before - len(frame)

    stats.final_rows = len(frame)
    stats.label_counts = {k: int(v) for k, v in frame["sentiment"].value_counts().items()}

    logger.info(
        "Dataset: %d dòng -> %d sau làm sạch "
        "(rỗng %d, nhãn rác %d, mâu thuẫn nhãn %d, trùng nội dung %d); phân bố: %s",
        stats.total_rows, stats.final_rows, stats.dropped_empty, stats.dropped_bad_label,
        stats.dropped_conflict, stats.dropped_duplicate, stats.label_counts,
    )
    return frame, stats


def split_dataset(
    frame: pd.DataFrame, test_size: float = 0.2, seed: int = 42
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Chia train/test có PHÂN TẦNG theo nhãn.

    `stratify` là bắt buộc: lớp `neutral` chỉ chiếm ~10%, split ngẫu nhiên có thể cho
    tập test lệch nặng và metrics mất ý nghĩa.
    """
    from sklearn.model_selection import train_test_split  # import trễ: sklearn nặng

    # `stratify` cần mỗi lớp có >= 2 mẫu, nếu không sklearn ném lỗi khó hiểu.
    # Dữ liệu như vậy vốn đã không dùng được, nhưng đừng để nó sập với thông điệp mù mờ.
    smallest = frame["sentiment"].value_counts().min()
    stratify = frame["sentiment"] if smallest >= 2 else None
    if stratify is None:
        logger.warning(
            "Có lớp chỉ %d mẫu - chia ngẫu nhiên không phân tầng, metrics sẽ không đáng tin",
            int(smallest),
        )

    train_df, test_df = train_test_split(
        frame,
        test_size=test_size,
        random_state=seed,
        stratify=stratify,
    )
    return train_df.reset_index(drop=True), test_df.reset_index(drop=True)


def iter_untagged_files(input_dir: str) -> Iterator[tuple[Path, list[dict]]]:
    """Duyệt thư mục Excel CHƯA gán nhãn, yield theo TỪNG FILE.

    Generator theo file để không ôm hàng trăm nghìn dòng trong RAM cùng lúc.
    """
    directory = Path(input_dir).resolve()
    if not directory.is_dir():
        raise FileNotFoundError(f"Không thấy thư mục đầu vào: {directory}")

    for path in sorted(directory.glob("*.xlsx")):
        try:
            sheet = load_workbook(path, read_only=True).active
            rows = sheet.iter_rows(values_only=True)
            header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
            idx_text = _column_index(header, COL_TEXT)
            idx_id = _column_index(header, COL_ID)
            idx_link = header.index(COL_LINK) if COL_LINK in header else None
            idx_name = header.index(COL_NAME) if COL_NAME in header else None
        except (StopIteration, ValueError, OSError, KeyError):
            logger.warning("Bỏ qua file không đọc được: %s", path.name, exc_info=True)
            continue

        items: list[dict] = []
        for row in rows:
            text = row[idx_text] if idx_text < len(row) else None
            if not text or not str(text).strip():
                continue
            items.append(
                {
                    "link": row[idx_link] if idx_link is not None and idx_link < len(row) else None,
                    "name_item": row[idx_name] if idx_name is not None and idx_name < len(row) else None,
                    "comments_id": row[idx_id] if idx_id < len(row) else None,
                    "comments_content": str(text),
                }
            )
        yield path, items


def label_distribution(frame: pd.DataFrame) -> dict[str, int]:
    return {k: int(v) for k, v in frame["sentiment"].value_counts().items()}


def describe_split(train_df: pd.DataFrame, test_df: pd.DataFrame) -> dict:
    return {
        "train": len(train_df),
        "test": len(test_df),
        "train_labels": label_distribution(train_df),
        "test_labels": label_distribution(test_df),
    }
