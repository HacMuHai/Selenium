"""
Test pipeline train/evaluate/metadata và trang HTML.

Trang HTML phải TỰ CHỨA (không tải gì từ mạng) và phải ESCAPE nội dung comment -
nó được serve qua `/analyze/report`, nội dung do người ngoài viết nên đây là đường XSS thật.
"""
import json

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.analysis import metrics
from src.analysis.dataset import describe_split
from src.analysis.predictor import ModelNotTrained, Predictor
from src.analysis.report import render_csv, render_html
from src.analysis.trainer import Trainer, load_metadata

POSITIVE = ["máy dùng rất tốt", "sản phẩm tuyệt vời", "pin trâu màn đẹp",
            "giao hàng nhanh nhiệt tình", "rất hài lòng nên mua"]
NEGATIVE = ["máy quá tệ nhanh hỏng", "pin tụt rất nhanh", "loa bị rè khó nghe",
            "dịch vụ bảo hành tệ", "màn hình lỗi không lên"]
NEUTRAL = ["cho hỏi máy này mấy sim", "bên mình có bán ốp lưng không",
           "giá thay màn hình bao nhiêu", "shop có thu cũ không",
           "cách cài tiếng việt thế nào"]


def make_frame():
    rows = (
        [{"text": t, "sentiment": "positive"} for t in POSITIVE * 4]
        + [{"text": t, "sentiment": "negative"} for t in NEGATIVE * 4]
        + [{"text": t, "sentiment": "neutral"} for t in NEUTRAL * 4]
    )
    return pd.DataFrame(rows)


@pytest.fixture
def trained(tmp_path):
    frame = make_frame()
    trainer = Trainer(tmp_path / "models", seed=42)
    results = trainer.train(["nb", "svm"], frame)
    evaluations = trainer.evaluate(["nb", "svm"], frame)
    baseline = metrics.majority_baseline(list(frame["sentiment"]))
    trainer.write_metadata(
        _stats(frame), describe_split(frame, frame), results, evaluations, baseline, 0.2
    )
    return trainer, load_metadata(tmp_path / "models")


def _stats(frame):
    from src.analysis.dataset import CleaningStats

    stats = CleaningStats()
    stats.total_rows = len(frame)
    stats.final_rows = len(frame)
    stats.label_counts = {k: int(v) for k, v in frame["sentiment"].value_counts().items()}
    return stats


def test_metadata_du_khoa(trained):
    _, metadata = trained

    assert metadata["schema"] == 1
    assert metadata["trained_at"]
    assert set(metadata["models"]) == {"nb", "svm"}
    assert metadata["baseline"]["label"] in {"negative", "neutral", "positive"}
    assert metadata["dataset"]["total_rows"] == 60


def test_evaluate_khong_train_lai(trained, tmp_path):
    """`evaluate` chỉ nạp artifact; không được đụng tới dữ liệu train."""
    trainer, _ = trained
    out = trainer.evaluate(["nb"], make_frame())

    assert 0.0 <= out["nb"]["macro_f1"] <= 1.0
    assert "predict_seconds" in out["nb"]


def test_evaluate_luu_vi_du_doan_sai(tmp_path):
    frame = make_frame()
    trainer = Trainer(tmp_path / "m", seed=1)
    trainer.train(["nb"], frame)
    # cố tình đánh giá trên nhãn đảo ngược -> chắc chắn có lỗi để lấy mẫu
    flipped = frame.copy()
    flipped["sentiment"] = flipped["sentiment"].map(
        {"positive": "negative", "negative": "positive", "neutral": "neutral"}
    )
    out = trainer.evaluate(["nb"], flipped)

    samples = out["nb"]["errors_sample"]
    assert samples and len(samples) <= 10
    assert set(samples[0]) == {"text", "true", "pred"}


def test_load_metadata_chua_train(tmp_path):
    assert load_metadata(tmp_path) is None


# ---------- HTML report ----------

def test_html_tu_chua_khong_tai_gi_tu_mang(trained):
    _, metadata = trained
    page = render_html(metadata)

    assert "http://" not in page and "https://" not in page
    assert "<svg" in page
    assert "baseline" in page


def test_html_co_du_ten_model(trained):
    _, metadata = trained
    page = render_html(metadata)

    assert "nb" in page and "svm" in page


def test_html_escape_noi_dung_chong_xss(trained):
    """Nội dung comment do người ngoài viết -> phải escape trước khi nhúng vào HTML."""
    _, metadata = trained
    metadata["models"]["nb"]["metrics"]["errors_sample"] = [
        {"text": "<script>alert(1)</script>", "true": "positive", "pred": "negative"}
    ]
    metadata["models"]["nb"]["metrics"]["macro_f1"] = 1.0   # ép nb thành "best"
    page = render_html(metadata)

    assert "<script>alert(1)</script>" not in page
    assert "&lt;script&gt;" in page


def test_html_khong_co_model():
    page = render_html({"models": {}, "baseline": {}, "dataset": {}, "split": {}})
    assert "Chưa có model" in page


def test_csv_co_dong_baseline(trained):
    _, metadata = trained
    text = render_csv(metadata)
    lines = text.strip().splitlines()

    assert lines[0].startswith("model,macro_f1,accuracy")
    assert any(line.startswith("baseline(") for line in lines)


# ---------- Predictor ----------

def test_predictor_chua_train_bao_loi_ro_rang(tmp_path):
    with pytest.raises(ModelNotTrained, match="src.analyze train"):
        Predictor(tmp_path).load("nb")


def test_predictor_ten_model_sai(tmp_path):
    with pytest.raises(ValueError):
        Predictor(tmp_path).load("khongco")


def test_predictor_cache_tra_ve_cung_object(trained, tmp_path):
    trainer, _ = trained
    predictor = Predictor(trainer.models_dir)
    assert predictor.load("nb") is predictor.load("nb")


def test_predict_dir_sinh_excel_co_cot_sentiment(trained, tmp_path):
    trainer, _ = trained
    source = tmp_path / "in"
    source.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["link", "name_item", "comments_id", "comments_content"])
    sheet.append(["l1", "SP", "c1", "máy dùng rất tốt"])
    sheet.append(["l2", "SP", "c2", "máy quá tệ nhanh hỏng"])
    workbook.save(str(source / "a.xlsx"))

    files = Predictor(trainer.models_dir).predict_dir(str(source), str(tmp_path / "out"), "nb")

    assert len(files) == 1
    rows = list(load_workbook(files[0]).active.iter_rows(values_only=True))
    assert rows[0] == ("link", "name_item", "comments_id", "comments_content",
                       "sentiment", "sentiment_model")
    assert rows[1][4] in {"negative", "neutral", "positive"}
    assert rows[1][5] == "nb"


def test_predict_dir_chan_ghi_de_thu_muc_dau_vao(trained, tmp_path):
    trainer, _ = trained
    same = tmp_path / "same"
    same.mkdir()
    with pytest.raises(ValueError, match="khác thư mục đầu vào"):
        Predictor(trainer.models_dir).predict_dir(str(same), str(same), "nb")
