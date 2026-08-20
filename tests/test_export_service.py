"""
Test ExportService qua InMemoryProductRepository - không cần Mongo, không mock openpyxl.
File sinh ra được đọc lại bằng openpyxl để kiểm nội dung thật.
"""
import pytest
from openpyxl import load_workbook

from src.repositories.memory_repository import InMemoryProductRepository
from src.services.export_service import HEADER, ExportService

from tests.conftest import make_product


def build_repo(product_count: int, comments_per_product: int, site: str = ""):
    repo = InMemoryProductRepository()
    for p in range(product_count):
        comments = [
            {
                "id": f"p{p}c{c}",
                "name": f"user{c}",
                "content": f"noi dung {p}-{c}",
                "rating": c % 6,
            }
            for c in range(comments_per_product)
        ]
        product = make_product(f"San pham {p}", f"https://example.com/p{p}", comments)
        if site:
            product["site"] = site
        repo.insert_product(product)
    return repo


def test_chia_file_theo_max_rows(tmp_path):
    repo = build_repo(product_count=3, comments_per_product=2)  # 6 dòng
    service = ExportService(repo, str(tmp_path), max_rows_per_file=2)

    files = service.export()

    assert len(files) == 3
    assert all(path.exists() for path in files)
    # Link example.com không thuộc sàn nào -> gom vào thư mục "khac".
    assert sorted(p.name for p in files) == [
        "comments_export_1.xlsx",
        "comments_export_2.xlsx",
        "comments_export_3.xlsx",
    ]
    assert all(path.parent == tmp_path / "khac" for path in files)


def test_header_va_noi_dung(tmp_path):
    repo = build_repo(product_count=1, comments_per_product=2)
    service = ExportService(repo, str(tmp_path), max_rows_per_file=100)

    files = service.export()

    sheet = load_workbook(files[0]).active
    rows = list(sheet.iter_rows(values_only=True))
    assert list(rows[0]) == HEADER
    # Dòng đầu của product có link + tên; dòng sau để trống 2 cột đầu.
    assert rows[1][0] == "https://example.com/p0"
    assert rows[1][3] == "noi dung 0-0"
    assert rows[2][0] is None
    assert rows[2][3] == "noi dung 0-1"


def test_bo_qua_comment_rong(tmp_path):
    repo = InMemoryProductRepository()
    repo.insert_product(
        make_product(
            "P",
            "https://example.com/p",
            [
                {"id": "a", "name": "x", "content": "", "rating": 0},
                {"id": "b", "name": "y", "content": "co noi dung", "rating": 1},
            ],
        )
    )
    service = ExportService(repo, str(tmp_path), max_rows_per_file=100)

    files = service.export()

    sheet = load_workbook(files[0]).active
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 dòng
    assert rows[1][2] == "b"


def test_khong_co_du_lieu_thi_khong_tao_file(tmp_path):
    service = ExportService(InMemoryProductRepository(), str(tmp_path))

    assert service.export() == []
    assert list(tmp_path.rglob("*.xlsx")) == []


def test_max_rows_khong_hop_le():
    with pytest.raises(ValueError):
        ExportService(InMemoryProductRepository(), "/tmp/x", max_rows_per_file=0)


def test_memory_repository_khong_dedup_theo_db():
    """--no-db không đọc Mongo nên exists_by_link luôn False (có chủ đích)."""
    repo = build_repo(product_count=1, comments_per_product=1)

    assert repo.exists_by_link("https://example.com/p0") is False
    assert repo.count_products() == 1


def test_tach_thu_muc_theo_san(tmp_path):
    """Mỗi sàn một thư mục, số thứ tự file đếm riêng."""
    repo = InMemoryProductRepository()
    for site, link in (
        ("cellphones", "https://cellphones.com.vn/a.html"),
        ("fptshop", "https://fptshop.com.vn/b"),
        ("cellphones", "https://cellphones.com.vn/c.html"),
    ):
        product = make_product("P", link, [{"id": link, "name": "u", "content": "x", "rating": 5}])
        product["site"] = site
        repo.insert_product(product)

    files = ExportService(repo, str(tmp_path), max_rows_per_file=100).export()

    assert {p.parent.name for p in files} == {"cellphones", "fptshop"}
    assert (tmp_path / "cellphones" / "comments_export_1.xlsx").exists()
    assert (tmp_path / "fptshop" / "comments_export_1.xlsx").exists()
    assert len(files) == 2  # 2 product cellphones nằm chung 1 file


def test_san_suy_ra_tu_link_khi_thieu_field_site(tmp_path):
    """Document cũ (crawl trước khi có field `site`) vẫn về đúng thư mục."""
    repo = InMemoryProductRepository()
    repo.insert_product(
        make_product(
            "P",
            "https://www.thegioididong.com/dtdd/x",
            [{"id": "a", "name": "u", "content": "noi dung", "rating": 4}],
        )
    )

    files = ExportService(repo, str(tmp_path), max_rows_per_file=100).export()

    assert files[0].parent.name == "thegioididong"


def test_file_moi_mo_dau_bang_dong_co_link(tmp_path):
    """Product bị cắt sang file mới thì file mới phải có lại link + tên."""
    repo = build_repo(product_count=1, comments_per_product=4, site="cellphones")

    files = ExportService(repo, str(tmp_path), max_rows_per_file=2).export()

    assert len(files) == 2
    for path in files:
        rows = list(load_workbook(path).active.iter_rows(values_only=True))
        assert rows[1][0] == "https://example.com/p0"
        assert rows[1][1] == "San pham 0"


def test_cot_site_user_rating(tmp_path):
    repo = build_repo(product_count=1, comments_per_product=1, site="fptshop")

    files = ExportService(repo, str(tmp_path), max_rows_per_file=100).export()

    rows = list(load_workbook(files[0]).active.iter_rows(values_only=True))
    assert rows[0][4:] == ("site", "user_name", "rating")
    assert rows[1][4:] == ("fptshop", "user0", 0)
