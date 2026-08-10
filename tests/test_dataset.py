"""
Test tầng dữ liệu bằng file Excel THẬT tạo trong tmp_path (không mock openpyxl).

Assert quan trọng nhất: KHÔNG RÒ RỈ giữa train và test.
"""
import pytest
from openpyxl import Workbook, load_workbook

from src.analysis.dataset import (
    iter_untagged_files,
    load_tagged_dataset,
    split_dataset,
)

HEADER = ["link", "name_item", "comments_id", "comments_content", "tag", "tag_cu", "nguon_tag"]


def write_tagged(directory, name, rows, header=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header or HEADER)
    for row in rows:
        sheet.append(row)
    path = directory / name
    workbook.save(str(path))
    return path


@pytest.fixture
def tag_dir(tmp_path):
    """Cố tình cài sẵn mọi ca bẩn: trùng nội dung, mâu thuẫn nhãn, nhãn rác, dòng rỗng."""
    d = tmp_path / "tags"
    d.mkdir()
    write_tagged(d, "a.xlsx", [
        ["l1", "SP 1", "c1", "Máy dùng rất tốt", "positive", "neutral", "gan-lai"],
        ["l1", "SP 1", "c2", "Máy Dùng Rất Tốt!!!", "positive", "neutral", "gan-lai"],  # trùng
        ["l1", "SP 1", "c3", "Pin tụt quá nhanh", "negative", "negative", "giu-nguyen"],
        ["l1", "SP 1", "c4", "   ", "positive", "", "gan-lai"],                          # rỗng
        ["l1", "SP 1", "c5", "Cho hỏi máy này mấy sim", "neutral", "neutral", "gan-lai"],
    ])
    write_tagged(d, "b.xlsx", [
        ["l2", "SP 2", "c6", "Loa bị rè", "negative", "neutral", "gan-lai"],
        ["l2", "SP 2", "c7", "Giao hàng nhanh", "rác", "", "gan-lai"],                    # nhãn rác
        ["l2", "SP 2", "c8", "Sản phẩm ổn", "positive", "", "gan-lai"],
        ["l2", "SP 2", "c9", "Sản phẩm ổn", "negative", "", "gan-lai"],                   # mâu thuẫn
        ["l2", "SP 2", "c10", "Máy chạy êm", "positive", "", "gan-lai"],
        ["l2", "SP 2", "c11", "Màn hình vỡ", "negative", "", "gan-lai"],
    ])
    return d


def test_lam_sach_dem_dung_tung_loai(tag_dir):
    frame, stats = load_tagged_dataset(str(tag_dir))

    assert stats.total_rows == 11
    assert stats.dropped_empty == 1
    assert stats.dropped_bad_label == 1
    assert stats.dropped_conflict == 2      # "Sản phẩm ổn" bị gán 2 nhãn -> bỏ CẢ HAI
    assert stats.dropped_duplicate == 1     # "Máy dùng rất tốt" vs "Máy Dùng Rất Tốt!!!"
    assert stats.final_rows == len(frame) == 6


def test_khong_giu_lai_noi_dung_mau_thuan_nhan(tag_dir):
    frame, _ = load_tagged_dataset(str(tag_dir))
    assert not frame["norm_text"].str.contains("sản phẩm ổn").any()


def test_doc_dung_cot_tag_khong_nham_tag_cu(tag_dir):
    """`tag_cu` chỉ để truy vết, tuyệt đối không được dùng làm nhãn train."""
    frame, _ = load_tagged_dataset(str(tag_dir))
    row = frame[frame["norm_text"].str.startswith("máy dùng rất tốt")].iloc[0]
    assert row["sentiment"] == "positive"   # tag_cu của dòng này là "neutral"


def test_thu_muc_khong_ton_tai(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_tagged_dataset(str(tmp_path / "khong-co"))


def test_thu_muc_rong(tmp_path):
    empty = tmp_path / "rong"
    empty.mkdir()
    with pytest.raises(ValueError):
        load_tagged_dataset(str(empty))


def test_file_thieu_cot_bi_bo_qua_khong_giet_pipeline(tag_dir):
    write_tagged(tag_dir, "hong.xlsx", [["x", "y"]], header=["link", "name_item"])
    frame, stats = load_tagged_dataset(str(tag_dir))

    assert "hong.xlsx" in stats.skipped_files
    assert len(frame) == 6      # các file lành vẫn đọc bình thường


def test_split_khong_ro_ri_du_lieu(tag_dir):
    """RỦI RO SỐ 1: cùng 1 comment nằm ở cả train lẫn test -> điểm số ảo."""
    frame, _ = load_tagged_dataset(str(tag_dir))
    train_df, test_df = split_dataset(frame, test_size=0.5, seed=42)

    assert set(train_df["norm_text"]) & set(test_df["norm_text"]) == set()
    assert len(train_df) + len(test_df) == len(frame)


def test_split_lop_chi_1_mau_khong_lam_sap(tag_dir):
    """`stratify` của sklearn cần >=2 mẫu mỗi lớp; fixture có đúng 1 dòng neutral."""
    frame, _ = load_tagged_dataset(str(tag_dir))
    assert (frame["sentiment"] == "neutral").sum() == 1

    train_df, test_df = split_dataset(frame, test_size=0.5, seed=42)
    assert len(train_df) + len(test_df) == len(frame)


def test_split_phan_tang_giu_ty_le_lop():
    import pandas as pd

    frame = pd.DataFrame(
        {
            "text": [f"cau so {i}" for i in range(300)],
            "norm_text": [f"cau so {i}" for i in range(300)],
            "sentiment": ["negative"] * 150 + ["positive"] * 120 + ["neutral"] * 30,
            "comment_id": [str(i) for i in range(300)],
        }
    )
    train_df, test_df = split_dataset(frame, test_size=0.2, seed=42)

    for label in ("negative", "positive", "neutral"):
        p_train = (train_df["sentiment"] == label).mean()
        p_test = (test_df["sentiment"] == label).mean()
        assert abs(p_train - p_test) < 0.02


def test_iter_untagged_files(tmp_path):
    d = tmp_path / "untagged"
    d.mkdir()
    for i in range(3):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["link", "name_item", "comments_id", "comments_content"])
        sheet.append([f"l{i}", f"SP {i}", f"c{i}", f"nội dung {i}"])
        sheet.append([None, None, f"c{i}x", "   "])       # dòng rỗng -> bỏ
        workbook.save(str(d / f"f{i}.xlsx"))

    batches = list(iter_untagged_files(str(d)))

    assert len(batches) == 3
    assert all(len(rows) == 1 for _, rows in batches)
    assert batches[0][1][0]["comments_content"] == "nội dung 0"
